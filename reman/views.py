from concurrent.futures import process
from django.db.models.fields import CharField
from django.http.response import HttpResponse, JsonResponse
from django.db.models.functions import Cast, Extract, datetime
from django.db import connection
from django.shortcuts import render
from .models import CC_material, FG_material, Midstep_material, Operation, Order, Component, Pseudo_R_material, R_material, RR_disassembly, Yfcorer, RR_sorting, Q3, OrderDates
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q, Count, Min, Max
import csv
import json
import datetime
from openpyxl import Workbook


class Material:
    def __init__(self, material_number):
        self.material_number = material_number
        self.quantity_as_input = 0
        self.quantity_as_output = 0
        self.parent_materials = []
        self.child_materials = []
        self.parent_consumption = {}
        self.child_production = {}
        self.orders = []
        self.renovation_rate = 0
        self.chain_position = 0
        self.chain = {}
        self.id = id(self)

    def add_parent_material(self, parent_material):
        if not parent_material in self.parent_materials:
            self.parent_materials.append(parent_material)

    def get_parent_materials(self):
        return self.parent_materials

    def add_order(self, order):
        self.orders.append(order)

    def add_child_material(self, child_material):
        if not child_material in self.child_materials:
            self.child_materials.append(child_material)

    def increase_quantity_as_input(self, input_quantity):
        self.quantity_as_input += input_quantity

    def increase_quantity_as_output(self, output_quantity):
        self.quantity_as_output += output_quantity

    def add_parent_consumption(self, parent_material, consumption):
        if not parent_material in self.parent_consumption:
            self.parent_consumption[parent_material] = consumption
        else:
            self.parent_consumption[parent_material] += consumption

    def get_parent_consumption(self, parent_material):
        return self.parent_consumption[parent_material]

    def add_child_production(self, child_material, production):
        if not child_material in self.child_production:
            self.child_production[child_material] = production
        else:
            self.child_production[child_material] += production

    def add_child_production_with_potential(self, child_material, production, potential):
        if not child_material in self.child_production:
            self.child_production[child_material] = [production, potential]
        else:
            self.child_production[child_material][0] += production
            self.child_production[child_material][1] += potential

    def get_child_production(self, child_material):
        return self.child_production[child_material]

    def get_renovation_rate(self):
        total_parent_consumption = 0
        for parent_material in self.parent_consumption:
            if 'FG' in parent_material.material_number or 'R' in parent_material.material_number:
                total_parent_consumption += self.parent_consumption[parent_material]
        try:
            return self.quantity_as_output / total_parent_consumption
        except:
            return 0

    def has_parent(self):
        return len(self.parent_materials) > 0

    def has_children(self):
        return len(self.child_materials) > 0

    def get_chain_position(self):
        examined_materials = [self]
        next_level_materials = []
        while len(examined_materials) > 0:
            for material in examined_materials:
                examined_materials.remove(material)
                next_level_materials.extend(material.get_parent_materials())
                
            if len(next_level_materials) > 0:
                self.chain_position += 1
                examined_materials = []
                examined_materials.extend(next_level_materials)
                next_level_materials = [] 

        return self.chain_position

    
    def get_components(self, components):
        self.components = []
        self.components.append(components)
    
    def show_chains_with_rate(self):
        #Zacinam rozebirat starou brzdu
        if len(self.parent_materials) == 0:
            list_of_quantity = list(Order.objects.filter(order_number__in = self.orders, material = self.material_number).values_list('order_quantity', 'confirmed_scrap'))
            order_quantity = sum(i for i, j in list_of_quantity)
            confirmed_scrap = sum(j for i, j in list_of_quantity)
            renovation_rate = (order_quantity - confirmed_scrap) / order_quantity
        elif self.chain_position == 1:
            list_of_quantity = list(Order.objects.filter(order_number__in = self.orders, material = self.material_number).values_list('order_quantity', 'confirmed_scrap'))
            order_quantity = sum(i for i, j in list_of_quantity)
            confirmed_scrap = sum(j for i, j in list_of_quantity)
            renovation_rate_p1 = (order_quantity - confirmed_scrap) / order_quantity
            list_of_quantity = list(Order.objects.filter(order_number__in = self.orders, material__in = self.parent_materials).values_list('order_quantity', 'confirmed_scrap'))
            order_quantity = sum(i for i, j in list_of_quantity)
            confirmed_scrap = sum(j for i, j in list_of_quantity)
            renovation_rate_p2 = (order_quantity - confirmed_scrap) / order_quantity
            renovation_rate = renovation_rate_p1 * renovation_rate_p2
        #Pokud ne, vezmu material, kde parent includuje CC (asi by se dalo vyresit chain_position == 1 - overit, zda plati)
        #Teoreticky cyklus while
        else:
            print('s')
        pass

        # projde zakazky, ktere material obsahuji (self.orders)
        # v zakazkach projde komponenty (order.components)

    def __eq__(self, other):
        return self.material_number == other.material_number

    def __repr__(self):
        return str(self.material_number)

    def __str__(self):
        return str(self.material_number)

    def __hash__(self):
        return hash(repr(self))


class MaterialCollection:
    def __init__(self):
        self.content = [] # materialy
        self.index_list = [] # materialove cisla
        self.material_chains = []

    def add_material(self, material):
        if not material.material_number in self.index_list:
            self.index_list.append(material.material_number)
            self.content.append(material)

    def get_or_create_material(self, material_number, *child_material):
        if child_material:
            child_material = self.content[self.index_list.index(child_material[0][0].material_number)]

        if material_number in self.index_list:
            material = self.content[self.index_list.index(material_number)]
            if child_material:
                for material_chain in self.material_chains:
                    for index, group in enumerate(material_chain):
                        if material in group:
                            if len(material_chain) > index + 1:
                                material_chain[index + 1].add(child_material)
                            else:
                                s = set()
                                s.add(child_material)
                                material_chain.append(s)
            return material
                
        else:
            material = Material(material_number)
            self.content.append(material)
            self.index_list.append(material_number)
            if child_material:
                s_p = set()
                s_ch = set()
                s_p.add(material)
                s_ch.add(child_material)
                self.material_chains.append([s_p, s_ch])
            else:
                s = set()
                s.add(material)
                self.material_chains.append([s])

            return material

    def get_fg_materials(self):
        return [material for material in self.content if 'FG' in material.material_number]

    def get_chains(self, material):
        return_list = []
        for material_chain in self.material_chains:
            for chain_element in material_chain:
                if material in chain_element:
                    return_list.append(material_chain)
        
        return return_list




@csrf_exempt
def orders(request=None):

    closed_nouns = ["TEUZ", "TECO"]

    if request.method == "POST":
        order_number = request.POST.get('order')
        type = request.POST.get('type')
        finish = request.POST.get('actual_finish')[:10] if request.POST.get('actual_finish') else None
        start = request.POST.get('start')[:10] if request.POST.get('start') else None
        closed = True if any(noun in request.POST.get('status') for noun in closed_nouns) else False
        dummy = True if int(request.POST.get('sum_value')) == 0 and type == "YRM2" else False
        wrong = True if (int(request.POST.get('target_value')) != int(request.POST.get('sum_value')) and type == "YRM2") else False
        obj, created = OrderDates.objects.update_or_create(order_number=order_number, defaults={'order_type': type, 'order_start_date': start, 'order_finish_date': finish, 'closed': closed, 'dummy': dummy, 'wrong': wrong})

    if request.method == "GET":
        if request.GET.get('which', 'open') == 'open':
            open_orders = list(OrderDates.objects.filter(closed=False, dummy=False, wrong=False).values_list('order_number', flat=True))
            return JsonResponse(open_orders, safe=False)
        if request.GET.get('which', 'open') == 'closed':
            closed_orders_without_date = OrderDates.objects.filter(closed=True, order_finish_date__isnull=True)
            for order in closed_orders_without_date:
                order.order_finish_date = order.order_start_date
                order.save()
            closed_orders = list(OrderDates.objects.filter(closed=True, dummy=False, wrong=False).annotate(month_closed=Extract('order_finish_date', 'month')).annotate(year_closed=Extract('order_finish_date', 'year')).filter(month_closed=3, year_closed=2022).values_list('order_number', flat=True))
            return JsonResponse(closed_orders, safe=False)

    return HttpResponse('okay')


def extract_materials(request=None):
    cc_materials = list(Component.objects.filter(pegged_requirement__startswith="CC").distinct('pegged_requirement').values_list('pegged_requirement', flat=True))

    for material in cc_materials:
        cc_material, created = CC_material.objects.get_or_create(material_number=material)
        fg_materials = list(Component.objects.filter(pegged_requirement=material).exclude(material=material).distinct('material').values_list('material', flat=True))

        for material in fg_materials:
            fg_material, created = FG_material.objects.get_or_create(material_number=material, cc_material=cc_material)
            if created:
                cc_material.new = True
                cc_material.save()
    
    r_materials = list(Component.objects.filter(Q(pegged_requirement__endswith="R") | Q(pegged_requirement__endswith="R2")).distinct('material').values_list('pegged_requirement', flat=True)) # muze byt R4

    # tady pridat r1/r3 fg1/fg3 

    for material in r_materials:
        r_material, created = R_material.objects.get_or_create(material_number=material)

    components = Component.objects.all()

    for component in components:
        if CC_material.objects.filter(material_number=component.pegged_requirement).exists():
            component.pegged_requirement_object_cc_material = CC_material.objects.get(material_number=component.pegged_requirement)
        if R_material.objects.filter(material_number=component.pegged_requirement).exists():
            component.pegged_requirement_object_r_material = R_material.objects.get(material_number=component.pegged_requirement)
        if FG_material.objects.filter(material_number=component.material).exists():
            component.material_object_fg_material.add(FG_material.objects.filter(material_number=component.material).first())
        if R_material.objects.filter(material_number=component.material).exists():
            component.material_r_material_object = R_material.objects.get(material_number=component.material)
        component.save()

    return HttpResponse('okay')


@csrf_exempt
def r_materials(request):
    if request.method == "GET":
        return JsonResponse(list(R_material.objects.all().values_list('material_number', flat=True).distinct('material_number')), safe=False)

    if request.method == "POST":
        
        data = json.loads(request.POST['materials'])
    
        for r_material, fg_materials in data.items():
            r_material_object = R_material.objects.get(material_number=r_material)
            for fg_material in fg_materials:
                if FG_material.objects.filter(material_number = fg_material).exists():
                    r_material_object.fg_material.add(fg_material_object)
                else:
                    fg_material_object, created = FG_material.objects.get_or_create(material_number=fg_material)
                    r_material_object.fg_material.add(fg_material_object)
                r_material_object.save()

    return HttpResponse('okay')


@csrf_exempt
def q3(request):

    if request.method == "GET":
        order_numbers = list(Order.objects.all().values_list('order_number', flat=True))

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="orders.csv"'

        writer = csv.writer(response)

        for order_number in order_numbers:
            writer.writerow([order_number])

        return response


    if request.method == "POST":
        file = request.FILES['file']

        for idx, line in enumerate(file):
            if idx > 0:
                line_text = line.decode().split(';')
            
                q3, created = Q3.objects.get_or_create(order_number=line_text[1], damage_code=line_text[5], defective_quantity=float(line_text[2]), damage_description=line_text[4], assembly=line_text[6].rstrip())

                if R_material.objects.filter(material_number=line_text[3]).exists():
                    r_material = R_material.objects.get(material_number=line_text[3])
                    q3.r_material.add(r_material)
                elif FG_material.objects.filter(material_number=line_text[3]).exists():
                    fg_material = FG_material.objects.get(material_number=line_text[3])
                    q3.fg_material.add(fg_material)
                elif CC_material.objects.filter(material_number=line_text[3]).exists():
                    cc_material = CC_material.objects.get(material_number=line_text[3])
                    q3.cc_material.add(cc_material)
                
                if R_material.objects.filter(material_number=line_text[6]).exists():
                    r_material = R_material.objects.get(material_number=line_text[6])
                    q3.r_material.add(r_material)
                elif FG_material.objects.filter(material_number=line_text[6]).exists():
                    fg_material = FG_material.objects.get(material_number=line_text[6])
                    q3.fg_material.add(fg_material)
                elif CC_material.objects.filter(material_number=line_text[6]).exists():
                    cc_material = CC_material.objects.get(material_number=line_text[6])
                    q3.cc_material.add(cc_material)


        return HttpResponse('okay')


@csrf_exempt
def edit_materials(request):

    context = {}

    if request.method == "GET":
        cc_materials = CC_material.objects.all()

        context.update({
            'cc_materials': cc_materials,
        })

    if request.method == "POST":
        pass
    
    return render(request, 'edit_materials.html', context)


def report_orders(request):
    orders = Order.objects.all()

    context = {
        'orders': orders,
    }

    return render(request, 'report_orders.html', context)


@csrf_exempt
def visualize_cc(request):
    if 'period' in request.session:
        month, year = int(request.session['period'][0]), int(request.session['period'][1])
    else:
        month, year = (datetime.datetime.now().month - 2, datetime.datetime.now().year)

    month_beginning = datetime.datetime(year=year, month=month, day = 1)
    month_end = (month_beginning + datetime.timedelta(days=32)).replace(day = 1) - datetime.timedelta(days = 1)

    actual_orders = OrderDates.objects.filter(order_finish_date__gte=month_beginning, order_finish_date__lte=month_end).annotate(char_id=Cast('order_number', CharField())).values_list('char_id', flat=True)

    cc_materials = CC_material.objects.all().order_by('material_number')
    actual_month_fg_materials = list(Component.objects.filter(order_id__in=actual_orders, material__contains='FG').values_list('material', flat=True))

    context = {
        'cc_materials': cc_materials,
        'actual_fg_materials': actual_month_fg_materials,
    }

    if request.method == "POST":
        data = json.loads(request.body)

        if data['purpose'] == 'cc_material':
            cc_material = CC_material.objects.get(material_number=data['cc_material'])
            cc_material.disabled = data['disabled']
            cc_material.save()

        if data['purpose'] == 'cc_material_new':
            cc_material = CC_material.objects.get(material_number=data['cc_material'])
            cc_material.new = data['new']
            cc_material.save()

        if data['purpose'] == 'multiplicator':
            CC_material_object = CC_material.objects.get(material_number=data['cc_material'])
            fg_material = FG_material.objects.get(material_number=data['fg_material'], cc_material=CC_material_object)
            fg_material.multiplicator = data['multiplicator']
            fg_material.save()

        if data['purpose'] == 'summary-group':
            cc_material = CC_material.objects.get(material_number=data['cc_material'])
            cc_material.summary_group = data['value']
            cc_material.save()

        if data['purpose'] == 'sorting-group':
            cc_material = CC_material.objects.get(material_number=data['cc_material'])
            cc_material.sorting_group = data['value']
            cc_material.save()

        if data['purpose'] == 'brand':
            cc_material = CC_material.objects.get(material_number=data['cc_material'])
            cc_material.brand = data['value']
            cc_material.save()


    return render(request, 'visualize_cc.html', context)


@csrf_exempt
def visualize_fg(request):
    fg_materials = FG_material.objects.all().order_by('material_number').exclude(cc_material__isnull=True).distinct('material_number')

    context = {
        'fg_materials': fg_materials,
    }

    if request.method == "POST":
        data = json.loads(request.body)

        if data['purpose'] == 'fg_material':
            fg_material = FG_material.objects.filter(material_number=data['fg_material'])
            for item in fg_material:
                item.disabled = data['disabled']
                item.save()

        if data['purpose'] == 'category':
            fg_materials = FG_material.objects.filter(material_number=data['fg_material'])
            fg_materials.update(category=data['category'])


    return render(request, 'visualize_fg.html', context)


@csrf_exempt
def visualize_yfcorer(request=None):

    yfcorers_without_cc = Yfcorer.objects.filter(cc_material__isnull=True)

    for yfcore in yfcorers_without_cc:
        cc_material, created = CC_material.objects.get_or_create(material_number=yfcore.raw_cc_material)
        yfcore.cc_material_id = cc_material
        yfcore.save()
    
    yfcorers = Yfcorer.objects.all().order_by('-change_date')

    context = {
        'yfcorer': yfcorers,
    }

    return render(request, 'visualize_yfcorer.html', context)


@csrf_exempt
def period_setter(request):

    period = json.loads(request.body)

    request.session['period'] = (period['month'], period['year'])

    return HttpResponse('okay')


@csrf_exempt
def report_yfcorer(request=None):
    
    if request.method == "GET":

        if 'period' in request.session:
            month, year = request.session['period'][0], request.session['period'][1]
        else:
            month, year = (datetime.datetime.now().month - 2, datetime.datetime.now().year)
        
        if 'recalc' in request.GET:

            yfcorers = Yfcorer.objects.exclude(cc_material__material_number__isnull=True).annotate(month=Extract('change_date', 'month')).exclude(cc_material__disabled=True).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).exclude(cc_material__disabled=True).values('cc_material__summary_group', 'cc_material__sorting_group', 'cc_material__material_number').annotate(good_sum=Count('warehouse_flag', filter=Q(warehouse_flag=1))).annotate(bad_sum=Count('warehouse_flag', filter=Q(warehouse_flag=3)))

            summary_groups_results = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).exclude(cc_material__disabled=True).values('cc_material__summary_group').annotate(good_sum=Count('warehouse_flag', filter=Q(warehouse_flag=1) & ~Q(damage_1='G-Competitor Product'))).annotate(bad_sum=Count('warehouse_flag', filter=Q(warehouse_flag=3)))

            sorting_groups_results = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).exclude(cc_material__disabled=True).values('cc_material__sorting_group').annotate(good_sum=Count('warehouse_flag', filter=Q(warehouse_flag=1) & ~Q(damage_1='G-Competitor Product'))).annotate(bad_sum=Count('warehouse_flag', filter=Q(warehouse_flag=3)))

            fg_categories = FG_material.objects.values_list('category', flat=True)

            data = {}
            
            for yfcore in yfcorers:
                cc_material_object = CC_material.objects.get(material_number=yfcore['cc_material__material_number'])
                data[yfcore['cc_material__material_number']] = {}
                data[yfcore['cc_material__material_number']]['good_sum'] = yfcore['good_sum']
                data[yfcore['cc_material__material_number']]['bad_sum'] = yfcore['bad_sum']
                data[yfcore['cc_material__material_number']]['brand'] = cc_material_object.brand
                data[yfcore['cc_material__material_number']]['summary_group'] = yfcore['cc_material__summary_group']
                data[yfcore['cc_material__material_number']]['summary_group_result'] = summary_groups_results.get(cc_material__summary_group = yfcore['cc_material__summary_group'])['good_sum'] / (summary_groups_results.get(cc_material__summary_group = yfcore['cc_material__summary_group'])['good_sum'] + summary_groups_results.get(cc_material__summary_group = yfcore['cc_material__summary_group'])['bad_sum'])
                data[yfcore['cc_material__material_number']]['sorting_group'] = yfcore['cc_material__sorting_group']
                data[yfcore['cc_material__material_number']]['sorting_group_results'] = sorting_groups_results.get(cc_material__sorting_group = yfcore['cc_material__sorting_group'])['good_sum'] / (sorting_groups_results.get(cc_material__sorting_group = yfcore['cc_material__sorting_group'])['good_sum'] + sorting_groups_results.get(cc_material__sorting_group = yfcore['cc_material__sorting_group'])['bad_sum'])
                RR_sorting.objects.update_or_create(cc_material=cc_material_object, month=month, year=year, defaults = {'brand': cc_material_object.brand, 'summary_group': yfcore['cc_material__summary_group'], 'summary_group_result': data[yfcore['cc_material__material_number']]['summary_group_result'], 'sorting_group': yfcore['cc_material__sorting_group'], 'sorting_group_result': data[yfcore['cc_material__material_number']]['sorting_group_results']})

            context = {
                'data': data,
                'month': month,
                'year': year,
            }

        else:

            raw_data = RR_sorting.objects.filter(month=month, year=year).order_by('summary_group', 'sorting_group')

            data = {}

            for line in raw_data:
                
                data[line.cc_material] = {}
                data[line.cc_material]['brand'] = line.brand
                data[line.cc_material]['summary_group'] = line.summary_group
                data[line.cc_material]['summary_group_result'] = line.summary_group_result
                data[line.cc_material]['sorting_group'] = line.sorting_group
                data[line.cc_material]['sorting_group_results'] = line.sorting_group_result

            context = {
                'data': data,
                'month': month,
                'year': year,
            }

        return render(request, 'report_yfcorer.html', context)

    if request.method == "POST":

        if 'period' in request.session:
            month, year = request.session['period'][0], request.session['period'][1]
        else:
            month, year = (datetime.datetime.now().month - 2, datetime.datetime.now().year)

        data = json.loads(request.body)

        if data['purpose'] == 'summary-group':

            yfcorers = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).filter(cc_material__summary_group=data['value']).values('coc_description', 'change_date', 'warehouse_flag', 'cc_material_id', 'core_group', 'damage_1', 'damage_2', 'damage_3', 'damage_4', 'damage_5')

        if data['purpose'] == 'brand':

            yfcorers = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).filter(cc_material__brand=data['value']).values('coc_description', 'change_date', 'warehouse_flag', 'cc_material_id', 'core_group', 'damage_1', 'damage_2', 'damage_3', 'damage_4', 'damage_5')

        if data['purpose'] == 'cc-material':
            yfcorers = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).filter(cc_material__material_number=data['value']).values('coc_description', 'change_date', 'warehouse_flag', 'cc_material_id', 'core_group', 'damage_1', 'damage_2', 'damage_3', 'damage_4', 'damage_5')
            
        if data['purpose'] == 'sorting-group':
            yfcorers = Yfcorer.objects.annotate(month=Extract('change_date', 'month')).annotate(year=Extract('change_date', 'year')).filter(month=month, year=year).filter(cc_material__sorting_group=data['value']).values('coc_description', 'change_date', 'warehouse_flag', 'cc_material_id', 'core_group', 'damage_1', 'damage_2', 'damage_3', 'damage_4', 'damage_5')

        if data['purpose'] == 'rr-sorting-group':
            print('tady Peter doplni - 368')
            #<td class="yfcorer-data-td" data-purpose="rr-sorting-group" data-value="{{values.sorting_group_results}}">{{values.sorting_group_results|to_percentage:"0"}}</td>
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = f"attachment; filename=yfcorer-{data['value']}.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'YFCORER'

        columns = [
            'COC',
            'Change date',
            'Warehouse flag',
            'CC material',
            'Core group',
            'Damage 1',
            'Damage 2',
            'Damage 3',
            'Damage 4',
            'Damage 5',
        ]

        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            
        for yfcore in yfcorers:
            row_num += 1

            row = [
                yfcore['coc_description'],
                yfcore['change_date'],
                yfcore['warehouse_flag'],
                yfcore['cc_material_id'],
                yfcore['core_group'],
                yfcore['damage_1'],
                yfcore['damage_2'],
                yfcore['damage_3'],
                yfcore['damage_4'],
                yfcore['damage_5'],
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


@csrf_exempt
def report_disassembly(request):
    
    if request.method == "GET":

        if 'period' in request.session:
            month, year = request.session['period'][0], request.session['period'][1]
        else:
            month, year = (datetime.datetime.now().month - 2, datetime.datetime.now().year)

        if 'recalc' in request.GET:

            month_orders = OrderDates.objects.annotate(finish_month=Extract('order_finish_date', 'month')).filter(finish_month=month).values_list('order_number', flat=True)

            components = Component.objects.filter(order__order_type="YRM4").filter(order__in=month_orders).exclude(material__startswith='CC').exclude(pegged_requirement_object_cc_material__isnull=True)

            data = {}

            for component in components:
                order = component.order
                component_material = component.material_object_fg_material.first()

                if not component_material in data:
                    data[component_material] = {}
                    data[component_material]['theoretical_gain'] = 0
                    data[component_material]['real_gain'] = 0
                    data[component_material]['regeneration_rate'] = 0
                    data[component_material]['category'] = component_material.category
                    data[component_material]['regeneration_rate_group'] = 0
        
                data[component_material]['theoretical_gain'] += order.order_quantity * component_material.multiplicator
                data[component_material]['real_gain'] += abs(component.quantity_withdrawn)
                data[component_material]['regeneration_rate'] = data[component_material]['real_gain'] / data[component_material]['theoretical_gain']
                
                RR_disassembly.objects.update_or_create(fg_material=component_material, month=month, year=year, defaults = {'material_result':  data[component_material]['regeneration_rate'], 'category': component_material.category, 'theoretical_gain': data[component_material]['theoretical_gain'], 'real_gain': data[component_material]['real_gain'], 'deffective_quantity': Q3.objects.filter(assembly=component_material.material_number, order_number__in=month_orders).aggregate(def_qty=Sum('defective_quantity'))['def_qty']})

            categories = FG_material.objects.distinct('category').values_list('category', flat=True)
            
            for category in categories:
                category_cc_materials = CC_material.objects.select_related().filter(fg_material__category=category).values_list('material_number', flat=True).distinct('material_number')
                cc_results = []
                for cc_material in category_cc_materials:
                    fg_materials = FG_material.objects.filter(cc_material=cc_material, category=category)
                    cc_results.append(sum([v['regeneration_rate'] for (k, v) in data.items() if (v['category'] == category and k in fg_materials and v['theoretical_gain'] > 0)]))

                cc_results = [item for item in cc_results if item > 0]
                category_result = sum(cc_results) / len(cc_results) if len(cc_results) > 0 else 0
                RR_disassembly.objects.filter(category=category, month=month, year=year).update(category_result=category_result)
                    
                for fg_material, values in data.items():
                    if values['category'] == category:
                        values['regeneration_rate_group'] = category_result

                sorted_data = dict(sorted(data.items(), key=lambda x:((x[0].cc_material.summary_group if x[0].cc_material.summary_group else ''), (x[0].category if x[0].category else ''))))
                
            context = {
                'data': data,
                'month': month,
                'year': year,
            }

        else:
            raw_data = RR_disassembly.objects.filter(month=month, year=year)
            data = {}

            for line in raw_data:
                data[line.fg_material] = {}
                data[line.fg_material]['regeneration_rate'] = line.material_result
                data[line.fg_material]['regeneration_rate_group'] = line.category_result
                data[line.fg_material]['theoretical_gain'] = line.theoretical_gain
                data[line.fg_material]['real_gain'] = line.real_gain
                data[line.fg_material]['deffective_quantity'] = line.deffective_quantity


            sorted_data = dict(sorted(data.items(), key=lambda x: ((x[0].cc_material.summary_group if x[0].cc_material.summary_group else ''), (x[0].category if x[0].category else ''))))

            context = {
                'data': sorted_data,
                'month': month,
                'year': year,
            }



        return render(request, 'report_disassembly.html', context)


    if request.method == "POST":

        data = json.loads(request.body)

        if data['purpose'] == 'category':

            category_materials = FG_material.objects.filter(category=data['value']).values_list('material_number', flat=True)

            components = Component.objects.filter(material__in=category_materials).values('workcenter','requirement_date','requirement_time','order', 'order__order_quantity', 'order__quantity_delivered', 'pegged_requirement','material','material_description','requirement_quantity','quantity_withdrawn','open_quantity','storage_type','storage_bin','shortage','missing_part')

        if data['purpose'] == 'fg-material':

            components = Component.objects.filter(material=data['value']).values('workcenter','requirement_date','requirement_time','order', 'order__order_quantity', 'order__quantity_delivered', 'pegged_requirement','material','material_description','requirement_quantity','quantity_withdrawn','open_quantity','storage_type','storage_bin','shortage','missing_part')

        if data['purpose'] == 'rr-disassembly':

            print('tady Peter doplni - 524')
            #<td class="disassembly-data-td" data-purpose="rr-disassembly" data-value="{{fg_values.regeneration_rate}}">{{fg_values.regeneration_rate|to_percentage:"2"}}</td>

        if data['purpose'] == 'cc-material':

            fg_materials = FG_material.objects.filter(cc_material=data['value']).values_list('material_number', flat=True)

            components = Component.objects.filter(material__in=fg_materials).values('workcenter','requirement_date','requirement_time','order', 'order__order_quantity', 'order__quantity_delivered', 'pegged_requirement','material','material_description','requirement_quantity','quantity_withdrawn','open_quantity','storage_type','storage_bin','shortage','missing_part')

        if data['purpose'] == 'rr-disassembly-group':

            print('tady Peter doplni - 535')
            #<td class="disassembly-data-td" data-purpose="rr-disassembly-group" data-value="{{fg_values.regeneration_rate_group}}">{{fg_values.regeneration_rate_group|to_percentage:"2"}}</td>

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = f"attachment; filename=disassembly-{data['value']}.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Components'

        columns = [
            'Workcenter',
            'Requirement date',
            'Requirement time',
            'Order',
            'Order quantity',
            'Order quantity delivered',
            'Pegged requirement',
            'Material',
            'Material description',
            'Requirement quantity',
            'Quantity withdrawn',
            'Storage type',
            'Storage bin',
            'Shortage',
            'Missing part',
        ]

        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            
        for component in components:
            row_num += 1

            row = [
                component['workcenter'], 
                component['requirement_date'], 
                component['requirement_time'], 
                component['order'],
                component['order__order_quantity'],
                component['order__quantity_delivered'], 
                component['pegged_requirement'], 
                component['material'], 
                component['material_description'], 
                component['requirement_quantity'], 
                component['quantity_withdrawn'], 
                component['open_quantity'], 
                component['storage_type'], 
                component['storage_bin'], 
                component['shortage'], 
                component['missing_part']
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


@csrf_exempt
def report_recondition(request):

    if request.method == "GET":

        if 'period' in request.session:
            month, year = request.session['period'][0], request.session['period'][1]
        else:
            month, year = (datetime.datetime.now().month - 2, datetime.datetime.now().year)

        month_orders = OrderDates.objects.annotate(finish_month=Extract('order_finish_date', 'month'), finish_year=Extract('order_finish_date', 'year')).filter(finish_month=month, finish_year=year).values_list('order_number', flat=True)

        r_materials = R_material.objects.all().values_list('material_number', flat=True).distinct()
        r_orders = Order.objects.filter(material__in=r_materials, order_number__in=month_orders).values_list('order_number', flat=True)
        r_operations = list(Operation.objects.filter(order__in=r_orders).values('order__material', 'order_id').annotate(final_yield=Min('confirmed_yield'), operation_qty=Max('operation_quantity')).values('order__material', 'final_yield', 'operation_qty'))

        results = {}

        for value_dict in r_operations:
            r_material = R_material.objects.get(material_number=value_dict['order__material'])
            if r_material not in results:
                results[r_material] = {}
                results[r_material]['yield'] = 0
                results[r_material]['operation_qty'] = 0
                results[r_material]['fg_materials'] = r_material.fg_material.filter(material_number__icontains='FG')
                results[r_material]['cc_materials'] = [fg_material.cc_material.material_number for fg_material in results[r_material]['fg_materials']]
            results[r_material]['yield'] += value_dict['final_yield']
            results[r_material]['operation_qty'] += value_dict['operation_qty']
            results[r_material]['recondition_rate'] = results[r_material]['yield'] / results[r_material]['operation_qty']


        with connection.cursor() as cursor:
            cursor.execute(f"WITH j_view AS (SELECT material, order_quantity, quantity_delivered, reman_order.order_number, SUM(reman_q3.defective_quantity) AS calc_defective_quantity FROM reman_order LEFT JOIN reman_q3 ON reman_q3.order_number = reman_order.order_number WHERE reman_order.order_type = 'YRM2' AND reman_order.order_number IN (SELECT order_number FROM reman_orderdates WHERE EXTRACT(MONTH FROM order_finish_date) = {month} AND EXTRACT(YEAR FROM order_finish_date) = {year}) GROUP BY reman_order.material, reman_order.order_number, reman_order.order_quantity, quantity_delivered) SELECT material, SUM(order_quantity) AS order_qty, SUM(quantity_delivered) AS qty_delivered, SUM(calc_defective_quantity) AS defective_qty FROM j_view GROUP BY material")
            rows = cursor.fetchall()

            q3_data = {}

            for row in rows:
                if R_material.objects.filter(material_number=row[0]).exists():
                    q3_r_material = R_material.objects.get(material_number=row[0])
                    q3_data[q3_r_material] = {}
                    q3_data[q3_r_material]['order_quantity'] = row[1]
                    q3_data[q3_r_material]['quantity_delivered'] = row[2]
                    q3_data[q3_r_material]['defective_quantity'] = row[3]
                    q3_data[q3_r_material]['difference'] = row[1] - row[2] - (row[3] or 0)

        for r_material in results:
            if r_material in q3_data:
                results[r_material]['order_quantity'] = q3_data[r_material]['order_quantity']
                results[r_material]['quantity_delivered'] = q3_data[r_material]['quantity_delivered']
                results[r_material]['defective_quantity'] = q3_data[r_material]['defective_quantity']
                results[r_material]['difference'] = q3_data[r_material]['difference']

        context = {
            'data': results,
            'month': month,
            'year': year,
        }

        return render(request, 'report_recondition.html', context)

    if request.method == "POST":

        data = json.loads(request.body)

        if data['purpose'] == 'cc-material':
 
            print('tady Peter doplni - 666')
            #<td class="recondition-data-td" data-purpose="cc-material" data-value="{{values.cc_materials}}">{{values.cc_materials|list_to_string|safe}}</td>
        
        if data['purpose'] == 'fg-material':

            print('tady Peter doplni - 671')
            #<td class="recondition-data-td" data-purpose="fg-material" data-value="{{values.fg_materials}}">{{values.fg_materials|queryset_to_list|list_to_string|safe}}</td>

        if data['purpose'] == 'r-material':

            r_orders = Order.objects.filter(material=data['value']).values_list('order_number', flat=True)

            operations = Operation.objects.filter(order__in=r_orders).values('material','order','activity','group','confirmation','workplace','operation_short_text','control_key','operation_quantity','confirmed_yield','scrap','rework','storage_location','system_status','actual_finish_date')

        if data['purpose'] == 'rr-recondition':

            print('tady Peter doplni 682')
            #<td class="recondition-data-td" data-purpose="rr-recondition" data-value="{{values.recondition_rate}}">{{values.recondition_rate|string_to_int|to_percentage:"0"}}</td>

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = f"attachment; filename=recondition-{data['value']}.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Operations'

        columns = [
            'Material',
            'Order',
            'Activity',
            'Group',
            'Confirmation',
            'Workplace',
            'Operation Short Text',
            'Control Key',
            'Operation qty',
            'Confirmed yield',
            'Scrap',
            'Rework',
            'Storage location',
            'System status',
            'Actual finish date',
        ]

        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            
        for operation in operations:
            row_num += 1

            row = [
                operation['material'], 
                operation['order'], 
                operation['activity'], 
                operation['group'], 
                operation['confirmation'], 
                operation['workplace'], 
                operation['operation_short_text'], 
                operation['control_key'], 
                operation['operation_quantity'], 
                operation['confirmed_yield'], 
                operation['scrap'], 
                operation['rework'], 
                operation['storage_location'], 
                operation['system_status'], 
                operation['actual_finish_date']
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


@csrf_exempt
def get_midstep_materials(request):

    if request.method == "GET":
        midstep_materials = FG_material.objects.all()
        
        midstep_materials = list(midstep_materials.filter(Q(material_number__endswith='FG1')|Q(material_number__endswith='FG3')).values_list('material_number', flat=True))

        return JsonResponse(midstep_materials, safe=False)
    
    if request.method == "POST":

        fg_material = request.POST.get('fg_material')
        pseudo_r_material = request.POST.get('pseudo_r_material')
        r_material = request.POST.get('r_material')

        if R_material.objects.filter(material_number = r_material).exists():

            pseudo_r_material_object, created = Pseudo_R_material.objects.get_or_create(material_number = pseudo_r_material, r_material = R_material.objects.get(material_number = r_material))
            midstep_material_object, created = Midstep_material.objects.get_or_create(pseudo_r_material = pseudo_r_material_object)
            midstep_material_object.fg_material.add(FG_material.objects.filter(material_number=fg_material).first())

        return HttpResponse('okay')


def order_story(request):

    orders = Order.objects.filter(order_number__in=OrderDates.objects.filter(order_finish_date__gte='2022-02-01').exclude(order_type__in=["YRM0", "YRM1"]).values_list('order_number', flat=True))

    material_collection = MaterialCollection()

    for order in orders:
    
        order_components = Component.objects.filter(order_id=order.order_number)
        order_material = material_collection.get_or_create_material(order.material)
        order_material.add_order(order.order_number)
        order_material.increase_quantity_as_output(order.quantity_delivered)

        if order.order_type == 'YRM2':    
            for component in order_components:
                parent_material = material_collection.get_or_create_material(component.material, [order_material])
                parent_material.add_child_material(order_material)
                parent_material.add_child_production(order_material, order.quantity_delivered)
                parent_material.increase_quantity_as_input(component.quantity_withdrawn)
                order_material.add_parent_material(parent_material)
                order_material.add_parent_consumption(parent_material, component.quantity_withdrawn)

        if order.order_type == 'YRM4':
            order_material.increase_quantity_as_input(order.order_quantity)
            for component in order_components:
                if 'FG' in component.material:
                    child_material = material_collection.get_or_create_material(component.material)
                    order_material.add_child_material(child_material)
                    order_material.add_child_production_with_potential(child_material, abs(component.quantity_withdrawn), order.order_quantity * FG_material.objects.get(material_number = component.material, cc_material=CC_material.objects.get(material_number = order.material)).multiplicator)
                    child_material.add_parent_material(order_material)
                    child_material.add_parent_consumption(order_material, order.order_quantity)
                    child_material.increase_quantity_as_output(abs(component.quantity_withdrawn))


    print(material_collection.get_chains(material_collection.get_or_create_material('K215182')))
    #K215182 = material_collection.get_chains(material_collection.get_or_create_material('K215182'))

    #for material in material_collection.content:
        #print(material)
        #.parent_consumption.items
        #.child_production.items



    context = {
        'materials': material_collection.content,
    }
        # if order.order_type == 'YRM4':
        #     print(f"Ze zakazky {order.order_number}, ktera je typu {order.order_type} jsme vytezili {[(abs(component.quantity_withdrawn), component.material) for component in order_components if component.material != component.pegged_requirement]}. Na vstupu bylo {order.material} a z nej jsme ocekavali {[(fg_material.material_number, fg_material.multiplicator * order.order_quantity) for fg_material in FG_material.objects.filter(cc_material=order.material)]}")


    return render(request, 'new_report.html', context)
